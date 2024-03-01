from openpyxl import load_workbook
from openpyxl import Workbook
import math
import random

# блок кода генерации случайных данных для задачи
# так как изначально было два кода, оставлю в неизмененном виде
# чтобы был понятен процесс работы
workbook = Workbook()
sheet = workbook.active

# 1 и 2 столбцы это координаты старта, 3 и 4 - конца, 5 - стоимость заказа
# 6 и 7 - координаты курьеров
# сто заказов, 20 курьеров
for i in range(1, 101):
    for j in range(1, 6):
        sheet.cell(row = i, column = j).value = round(random.uniform(0.00, 11000.00), 4)

# так как площадь г. Якутска 122 квадратных км, за пример был взят квадрат со сторонами 11 км

for i in range(1, 21):
    for j in range(6, 8):
        sheet.cell(row = i, column = j).value = round(random.uniform(0.00, 11000.00), 4)

workbook.save(filename="Coordinates.xlsx")

# блок кода с решением с заданными случайным образом данными

workbook = load_workbook(filename="Coordinates.xlsx")
sheet = workbook.active

# функция расстояния между двумя точками на плоскости
def dist(point1, point2):
    return math.sqrt((point1[0] - point2[0])**2 + (point1[1] - point2[1])**2)

# функция нахождения расстояния до каждого заказа от каждого другого заказа и каждого курьера (с учетом расстояния доставки)
def couriers_to_orders(couriers, orders):
    distance = {}
    for order in orders:
        distance[order['id']] = {}
        for order2 in orders:
            # так как мы учитываем расстояние доставки заказа, у нас получается подобие ориентированного графа
            if order['id'] != order2['id']:
                distance[order['id']][order2['id']] = dist(order['start'], order2['end']) + dist(order['start'], order['end'])
    for courier in couriers:
        for order in orders:
            distance[order['id']][courier] = dist(couriers[courier], order['start']) + dist(order['start'], order['end'])
    return distance

"""
основная функция для динамического решения
сначала я сделал решение многократным перебором 
где минимальное расстояние от курьера до точки запписывалось и цикл продолжался
потом я спросил у ChatGPT какой есть алгоритм для решения данной задачи
он выдал алгоритм Дейкстры, но в решении его даже не использовал
изучив другие алгоритмы на графы
я пришел к выводу, что тут нужно решать динамически, 
так как курьеры постоянно меняют свои координаты на конечные координаты заказа
с помощью функции couriers_to_orders программе не надо каждый раз рассчитывать расстояние между точками
а лишь обновлять координаты курьера и путь, который он прошел
а пройденный заказ просто удаляется
и функция повторяется, пока не кончатся невыполненные заказы
"""

def distribute(distance, distributed_orders):
    min_val = float('inf')
    min_from = ''
    min_to = -1
    for i in distance:
        for j in distance[i].keys():
            if isinstance(j, str):
                if min_val > distance[i][j]:
                    min_from = j
                    min_to = i
                    min_val = distance[i][j]
                    break
    distance.pop(min_to)
    for i in distance:
        del distance[i][min_to]
        distance[i][min_from] += min_val
    distributed_orders[min_from].append(min_to)
    if distance:
        distribute(distance, distributed_orders)
    else:
        return

#чтение из таблицы, сгенерированной ранее

n = 0
m = 0
r = 0
orders = []
couriers = {}

for row in sheet.iter_rows(values_only=True):
    m += 1
    orders.append({'id': m, 'start': [row[0], row[1]], 'end': [row[2], row[3]], 'cost': row[4]})
    if row[5]:
        n += 1
        couriers['courier'+ str(n)] = [row[5], row[6]]

distributed_orders = {}
for i in couriers:
    distributed_orders[i] = []

cours_to_ords = couriers_to_orders(couriers, orders)
distribute(cours_to_ords, distributed_orders)

# вывод ответов в консоль
for i in distributed_orders:
    print(i)
    print(distributed_orders[i])
