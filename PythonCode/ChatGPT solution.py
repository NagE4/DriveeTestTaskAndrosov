import heapq
import math
from openpyxl import load_workbook
from openpyxl import Workbook

# Функция для вычисления расстояния между двумя точками по координатам
def calculate_distance(point1, point2):
    return math.sqrt((point1[0] - point2[0])**2 + (point1[1] - point2[1])**2)

# Функция для нахождения кратчайшего пути между точками с использованием алгоритма Дейкстры
def shortest_path(graph, start, end):
    queue = [(0, start)]
    visited = set()

    while queue:
        cost, node = heapq.heappop(queue)
        if node not in visited:
            visited.add(node)

            if node == end:
                return cost

            for neighbor, weight in graph[node].items():
                heapq.heappush(queue, (cost + weight, neighbor))

    return float('inf')

# Функция для распределения заказов между курьерами с учетом приоритета скорости доставки
def distribute_orders(orders, couriers):
    courier_distances = {}
    for courier in couriers:
        courier_distances[courier] = {}

    for courier in couriers:
        courier_coordinates = couriers[courier]
        for order in orders:
            order_coordinates_from = order['from']
            order_coordinates_to = order['to']
            distance_from = calculate_distance(courier_coordinates, order_coordinates_from)
            distance_to = calculate_distance(order_coordinates_from, order_coordinates_to)
            total_distance = distance_from + distance_to
            courier_distances[courier][order['id']] = total_distance

    print(courier_distances)
    for order in orders:
        min_distance = float('inf')
        min_courier = None
        for courier in couriers:
            distance = courier_distances[courier][order['id']]
            if distance < min_distance:
                min_distance = distance
                min_courier = courier
        order['courier'] = min_courier

    return orders

workbook = load_workbook(filename="Coordinates.xlsx")
sheet = workbook.active

n = 0
m = 0
r = 0
orders = []
couriers = {}

for row in sheet.iter_rows(values_only=True):
    m += 1
    orders.append({'id': m, 'from': [row[0], row[1]], 'to': [row[2], row[3]]})
    if row[5]:
        n += 1
        couriers['courier'+ str(n)] = [row[5], row[6]]

# Пример входных данных


# Распределение заказов между курьерами
distributed_orders = distribute_orders(orders, couriers)
print(distributed_orders)