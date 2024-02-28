from openpyxl import Workbook
import random

workbook = Workbook()
sheet = workbook.active

# 1 и 2 столбцы это координаты старта, 3 и 4 - конца, 5 - стоимость заказа
# 6 и 7 - координаты курьеров
# сто заказов, 20 курьеров
for i in range(1,101):
    for j in range(1,6):
        sheet.cell(row = i, column = j).value = round(random.uniform(0.00, 11000.00), 4)

# так как площадь г. Якутска 122 квадратных км, за пример был взят квадрат со сторонами 11 км

for i in range(1,21):
    for j in range(6,8):
        sheet.cell(row = i, column = j).value = round(random.uniform(0.00, 11000.00), 4)

workbook.save(filename="Coordinates.xlsx")
